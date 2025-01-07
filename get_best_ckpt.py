import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 读取Excel文件中的数据
file_path = '/data1/jianglei/SAM-DiffSR/exp/result/checkpoints/sam_diffsr_df2k4x_caption_bs64/IQA-val-benchmark_loop-sam_diffsr_df2k4x_caption_bs64.xlsx'  # 替换为你的xlsx文件路径
df = pd.read_excel(file_path)

# 获取所有列名（假设第一列是 'exp'）
all_columns = df.columns[1:]

# 根据列名判断哪些是需要最大化的列，哪些是需要最小化的列
max_columns = [col for col in all_columns if 'psnr-Y' in col or 'ssim' in col]
min_columns = [col for col in all_columns if 'fid' in col]

# 找出每列的最大值和最小值
max_values = df[max_columns].max()
min_values = df[min_columns].min()

# 设置高亮样式（黄色填充）
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 统计每个实验的最优秀指标数，并记录最优秀的指标
max_count = []
best_metrics = []  # 用来记录每个实验的最优秀指标
highlight_cells = []  # 用来记录每个实验中最优秀的单元格位置

for i, row in df.iterrows():
    count = 0
    best_metric = []  # 记录当前实验的最优秀指标
    experiment_highlight_cells = []  # 当前实验的高亮单元格位置
    
    # 对于最大化指标，检查是否是最大值
    for col in max_columns:
        if row[col] == max_values[col]:
            count += 1
            best_metric.append(f"{col}: {row[col]} (Max)")
            experiment_highlight_cells.append((col, i + 2))  # 行数是i + 2，因为Excel索引从1开始，第一行是表头
    
    # 对于最小化指标，检查是否是最小值
    for col in min_columns:
        if row[col] == min_values[col]:
            count += 1
            best_metric.append(f"{col}: {row[col]} (Min)")
            experiment_highlight_cells.append((col, i + 2))  # 行数是i + 2

    max_count.append(count)
    best_metrics.append(best_metric)
    highlight_cells.append(experiment_highlight_cells)

# 将统计结果和最优秀指标添加到DataFrame
df['max_count'] = max_count
df['best_metrics'] = best_metrics

# 加载原Excel文件以进行高亮操作
wb = load_workbook(file_path)
ws = wb.active  # 默认选择活动表

# 高亮最优秀指标的单元格
for experiment_cells in highlight_cells:
    for col_name, row_idx in experiment_cells:
        col_idx = df.columns.get_loc(col_name) + 1  # 获取列的索引（Excel是1-based）
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = highlight_fill  # 应用高亮样式

# 保存修改后的文件
output_file = 'highlighted_output.xlsx'  # 输出的文件路径
wb.save(output_file)

# 输出结果
max_exp = df.loc[df['max_count'].idxmax()]
print(f"实验 {max_exp['exp']} 具有最多最优秀的指标，最优秀的指标数量为 {max_exp['max_count']}")
print("最优秀的指标如下：")
for metric in max_exp['best_metrics']:
    print(", ".join(metric))
